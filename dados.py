from openpyxl import load_workbook
wb = load_workbook('dimensionamento.xlsx')
worksheet = wb['Dados']

#Coleta de dados dos equipamentos

for j in range(8):
    j=6   
    rated_voltage = worksheet.cell(j, 2).value
    rated_power = worksheet.cell(j, 4).value
    system_type = worksheet.cell(j, 3).value
    rated_current = worksheet.cell(j, 5).value
    power_factor = worksheet.cell(j,6).value
    system_yield = worksheet.cell(j, 7).value
    length = worksheet.cell(j, 8).value
    installation_method = worksheet.cell(j, 9).value
    environmental_temperature = worksheet.cell(j, 10).value
    grouping_factor = worksheet.cell(j, 11).value
    soil_temperatue = worksheet.cell(j, 12).value
    overload_factor = worksheet.cell(j, 13).value
    soil_resistivity = worksheet.cell(j, 14).value
    permissible_voltage = worksheet.cell(j, 15).value
    resistance = worksheet.cell(j, 16).value
    reactance = worksheet.cell(j, 17).value
    cable_phase = worksheet.cell(j, 18).value
    insulation = worksheet.cell(j, 19).value
    short_circuit_current = worksheet.cell(j,20).value
    short_circuit_time = worksheet.cell(j,21).value
    insulation_time = worksheet.cell(j, 22).value
    normal_regime_time = worksheet.cell(j, 23).value

